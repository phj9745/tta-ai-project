"""Workflow helpers for Drive feature list spreadsheets."""
from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Optional, Sequence, Tuple

from fastapi import HTTPException

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]
from ..excel_templates.models import FEATURE_LIST_EXPECTED_HEADERS
from ..excel_templates.utils import AI_CSV_DELIMITER
from .naming import drive_name_matches, looks_like_header_row
from .templates import FEATURE_LIST_SHEET_CANDIDATES, FEATURE_LIST_START_ROW

__all__ = [
    "parse_feature_list_workbook",
    "build_feature_list_rows_csv",
    "prepare_feature_list_response",
]


def parse_feature_list_workbook(workbook_bytes: bytes) -> Tuple[str, int, List[str], List[Dict[str, str]]]:
    if load_workbook is None:  # pragma: no cover
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 필요합니다.")
    buffer = io.BytesIO(workbook_bytes)
    try:
        workbook = load_workbook(buffer, data_only=True)
    except Exception as exc:  # pragma: no cover - 안전망
        raise HTTPException(status_code=500, detail="엑셀 파일을 읽는 중 오류가 발생했습니다.") from exc

    headers = list(FEATURE_LIST_EXPECTED_HEADERS)
    extracted_rows: List[Dict[str, str]] = []
    sheet_title = ""
    start_row = FEATURE_LIST_START_ROW
    header_row_values: Optional[Sequence[Any]] = None
    column_map: Dict[str, int] = {}
    try:
        sheet = workbook.active
        selected_title = sheet.title
        for candidate in FEATURE_LIST_SHEET_CANDIDATES:
            matched = False
            for title in workbook.sheetnames:
                if drive_name_matches(title, candidate):
                    try:
                        sheet = workbook[title]
                        selected_title = sheet.title
                        matched = True
                        break
                    except KeyError:
                        continue
            if matched:
                break

        sheet_title = selected_title or ""
        max_col = max(len(headers), sheet.max_column or len(headers))
        header_row_index: Optional[int] = None
        first_data_row_index: Optional[int] = None
        for idx, row in enumerate(
            sheet.iter_rows(min_row=1, max_col=max_col, values_only=True),
            start=1,
        ):
            row_values: Sequence[Any] = row if isinstance(row, Sequence) else tuple()
            if not any(value is not None for value in row_values):
                continue

            if header_row_values is None:
                normalized = [
                    str(value).strip() if value is not None else ""
                    for value in row_values
                ]
                if looks_like_header_row(normalized, headers):
                    header_row_values = normalized
                    header_row_index = idx
                    continue

            if header_row_values is None:
                continue

            if first_data_row_index is None:
                first_data_row_index = idx

            if not column_map and header_row_values:
                for header_name in headers:
                    try:
                        column_index = header_row_values.index(header_name)
                        column_map[header_name] = column_index
                    except ValueError:
                        column_map[header_name] = headers.index(header_name)

            if looks_like_header_row(row_values, headers):
                continue

            row_data: Dict[str, str] = {}
            has_values = False
            for header_name in headers:
                column_index = column_map.get(header_name)
                cell_value = (
                    row_values[column_index]
                    if column_index is not None and column_index < len(row_values)
                    else None
                )
                text = "" if cell_value is None else str(cell_value).strip()
                if text:
                    has_values = True
                row_data[header_name] = text

            if not has_values:
                continue

            description = row_data.get("기능 설명", "")
            if not description:
                description = row_data.get("기능 개요", "")

            extracted_rows.append(
                {
                    "majorCategory": row_data.get("대분류", ""),
                    "middleCategory": row_data.get("중분류", ""),
                    "minorCategory": row_data.get("소분류", ""),
                    "featureDescription": description,
                }
            )

        if header_row_index is not None:
            start_row = header_row_index + 1
        if first_data_row_index is not None:
            start_row = first_data_row_index
    finally:
        workbook.close()

    if not sheet_title:
        sheet_title = "기능리스트"

    return sheet_title, start_row, headers, extracted_rows


def build_feature_list_rows_csv(rows: Sequence[Dict[str, str]]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=list(FEATURE_LIST_EXPECTED_HEADERS),
        lineterminator="\n",
        delimiter=AI_CSV_DELIMITER,
    )
    writer.writeheader()

    has_overview_column = "기능 개요" in FEATURE_LIST_EXPECTED_HEADERS

    for row in rows:
        major = str(row.get("majorCategory", "") or "").strip()
        middle = str(row.get("middleCategory", "") or "").strip()
        minor = str(row.get("minorCategory", "") or "").strip()
        description = str(row.get("featureDescription", "") or "").strip()

        if not any([major, middle, minor, description]):
            continue

        entry = {
            "대분류": major,
            "중분류": middle,
            "소분류": minor,
            "기능 설명": description,
        }
        if has_overview_column:
            entry["기능 개요"] = ""

        writer.writerow(entry)

    return output.getvalue()


def prepare_feature_list_response(
    *,
    file_id: str,
    file_name: str,
    sheet_name: str,
    start_row: int,
    headers: Sequence[str],
    rows: Sequence[Dict[str, str]],
    modified_time: Optional[str],
    project_overview: Optional[str],
) -> Dict[str, Any]:
    response: Dict[str, Any] = {
        "fileId": file_id,
        "fileName": file_name,
        "sheetName": sheet_name,
        "startRow": start_row,
        "headers": list(headers),
        "rows": list(rows),
        "modifiedTime": modified_time,
    }
    if project_overview is not None:
        response["projectOverview"] = project_overview
    return response
