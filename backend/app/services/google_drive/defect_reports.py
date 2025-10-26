import csv
import io
import logging
from functools import cmp_to_key
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from fastapi import HTTPException

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]

from ..excel_templates.models import (
    DEFECT_REPORT_EXPECTED_HEADERS,
    DEFECT_REPORT_START_ROW,
    DefectReportImage,
)
from ..excel_templates.utils import AI_CSV_DELIMITER
from .naming import drive_name_matches, looks_like_header_row

__all__ = [
    "DEFECT_REPORT_EXPECTED_HEADERS",
    "parse_defect_report_workbook",
    "build_defect_report_rows_csv",
    "prepare_defect_report_response",
    "normalize_defect_report_rows",
    "serialize_defect_report_images",
]

logger = logging.getLogger(__name__)

_DEFECT_SHEET_CANDIDATES: Tuple[str, ...] = (
    "결함리포트",
    "결함 리포트",
    "defect report",
    "defect_report",
)

_FIELD_KEY_MAP: Dict[str, str] = {
    "순번": "order",
    "시험환경(OS)": "environment",
    "시험환경 OS": "environment",
    "시험 환경": "environment",
    "시험환경": "environment",
    "결함요약": "summary",
    "결함 요약": "summary",
    "결함정도": "severity",
    "결함 정도": "severity",
    "발생빈도": "frequency",
    "발생 빈도": "frequency",
    "품질특성": "quality",
    "품질 특성": "quality",
    "결함 설명": "description",
    "업체 응답": "vendorResponse",
    "수정여부": "fixStatus",
    "수정 여부": "fixStatus",
    "비고": "note",
}

_COLUMN_HEADERS: Dict[str, str] = {
    "order": "순번",
    "environment": "시험환경(OS)",
    "summary": "결함요약",
    "severity": "결함정도",
    "frequency": "발생빈도",
    "quality": "품질특성",
    "description": "결함 설명",
    "vendorResponse": "업체 응답",
    "fixStatus": "수정여부",
    "note": "비고",
}

_DEFAULT_ROW: Dict[str, str] = {
    "order": "",
    "environment": "시험환경 모든 OS",
    "summary": "",
    "severity": "",
    "frequency": "",
    "quality": "",
    "description": "",
    "vendorResponse": "",
    "fixStatus": "",
    "note": "",
}

_QUALITY_ORDER: Tuple[str, ...] = (
    "기능적합성",
    "성능효율성",
    "호환성",
    "사용성",
    "신뢰성",
    "보안성",
    "유지보수성",
    "이식성",
    "일반적 요구사항",
)

_SEVERITY_ORDER: Dict[str, int] = {"H": 0, "M": 1, "L": 2}


def _normalize_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return text.strip()


def _normalize_key(key: str) -> str:
    return key.strip().lower().replace(" ", "")


def _quality_rank(value: str) -> int:
    text = value.strip()
    if not text:
        return len(_QUALITY_ORDER)
    for index, label in enumerate(_QUALITY_ORDER):
        if text == label:
            return index
    normalized = text.replace(" ", "")
    for index, label in enumerate(_QUALITY_ORDER):
        if normalized == label.replace(" ", ""):
            return index
    return len(_QUALITY_ORDER)


def _severity_rank(value: str) -> int:
    normalized = value.strip().upper()
    return _SEVERITY_ORDER.get(normalized, len(_SEVERITY_ORDER))


def _compare_rows(left: Dict[str, str], right: Dict[str, str]) -> int:
    quality_left = _quality_rank(left.get("quality", ""))
    quality_right = _quality_rank(right.get("quality", ""))
    if quality_left != quality_right:
        return -1 if quality_left < quality_right else 1

    severity_left = _severity_rank(left.get("severity", ""))
    severity_right = _severity_rank(right.get("severity", ""))
    if severity_left != severity_right:
        return -1 if severity_left < severity_right else 1

    env_left = left.get("environment", "")
    env_right = right.get("environment", "")
    if env_left != env_right:
        return -1 if env_left > env_right else 1

    summary_left = left.get("summary", "")
    summary_right = right.get("summary", "")
    if summary_left != summary_right:
        return -1 if summary_left > summary_right else 1

    description_left = left.get("description", "")
    description_right = right.get("description", "")
    if description_left != description_right:
        return -1 if description_left > description_right else 1

    return 0


def _detect_severity(value: str) -> str:
    normalized = value.strip()
    if not normalized:
        return ""
    upper = normalized.upper()
    if upper in {"H", "M", "L"}:
        return upper

    compact = upper.replace(" ", "")
    if any(token in compact for token in ["HIGH", "CRITICAL", "치명", "중대", "심각", "높음"]):
        return "H"
    if any(token in compact for token in ["MEDIUM", "중간", "보통", "일반"]):
        return "M"
    if any(token in compact for token in ["LOW", "경미", "낮음", "사소", "경미함", "미미"]):
        return "L"
    return normalized


def _detect_frequency(value: str) -> str:
    normalized = value.strip()
    if not normalized:
        return ""
    upper = normalized.upper()
    if upper in {"A", "R"}:
        return upper

    compact = upper.replace(" ", "")
    if any(token in compact for token in ["ALWAYS", "항상", "항시", "상시", "지속", "매번", "항구"]):
        return "A"
    if any(token in compact for token in ["INTERMITTENT", "SOMETIMES", "OCCASIONAL", "RARE", "간헐", "가끔", "드물", "재현", "비정기", "때때로", "조건부"]):
        return "R"
    return normalized


def _normalize_quality(value: str) -> str:
    normalized = value.strip()
    if not normalized:
        return ""
    compact = normalized.replace(" ", "")
    for label in _QUALITY_ORDER:
        if compact == label.replace(" ", ""):
            return label
    return normalized


def normalize_defect_record(entry: Mapping[str, Any]) -> Dict[str, str]:
    record = dict(_DEFAULT_ROW)

    for key in _COLUMN_HEADERS:
        if key in entry:
            record[key] = _normalize_value(entry.get(key))

    for header, key in _FIELD_KEY_MAP.items():
        if header not in entry:
            continue
        value = _normalize_value(entry.get(header))
        record[key] = value

    record["environment"] = "시험환경 모든 OS"
    record["vendorResponse"] = ""
    record["severity"] = _detect_severity(record.get("severity", ""))
    record["frequency"] = _detect_frequency(record.get("frequency", ""))
    record["quality"] = _normalize_quality(record.get("quality", ""))
    return record


def normalize_defect_report_rows(rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, str]]:
    normalized = [normalize_defect_record(row) for row in rows if row]
    normalized.sort(key=cmp_to_key(_compare_rows))

    for index, row in enumerate(normalized, start=1):
        row["order"] = str(index)
    return normalized


def parse_defect_report_workbook(workbook_bytes: bytes) -> Tuple[str, int, List[str], List[Dict[str, str]]]:
    if load_workbook is None:  # pragma: no cover
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 필요합니다.")

    buffer = io.BytesIO(workbook_bytes)
    try:
        workbook = load_workbook(buffer, data_only=True)
    except Exception as exc:  # pragma: no cover - safety net
        raise HTTPException(status_code=500, detail="엑셀 파일을 읽는 중 오류가 발생했습니다.") from exc

    headers = list(DEFECT_REPORT_EXPECTED_HEADERS)
    extracted_rows: List[Dict[str, str]] = []
    sheet_title = ""
    start_row = DEFECT_REPORT_START_ROW
    try:
        sheet = workbook.active
        selected_title = sheet.title
        for candidate in _DEFECT_SHEET_CANDIDATES:
            matched = False
            for title in workbook.sheetnames:
                if drive_name_matches(title, candidate):
                    try:
                        sheet = workbook[title]
                        selected_title = sheet.title
                        matched = True
                        break
                    except KeyError:
                        continue
            if matched:
                break

        sheet_title = selected_title or ""
        max_col = max(len(headers), sheet.max_column or len(headers))
        header_row_values: Optional[Sequence[Any]] = None
        header_row_index: Optional[int] = None
        column_map: Dict[str, int] = {}

        for idx, row in enumerate(
            sheet.iter_rows(min_row=1, max_col=max_col, values_only=True),
            start=1,
        ):
            row_values: Sequence[Any] = row if isinstance(row, Sequence) else tuple()
            if not any(value is not None for value in row_values):
                continue

            if header_row_values is None:
                normalized_values = [_normalize_value(value) for value in row_values]
                if looks_like_header_row(normalized_values, headers):
                    header_row_values = normalized_values
                    header_row_index = idx
                    for col_index, cell_value in enumerate(normalized_values):
                        normalized_key = _normalize_key(cell_value)
                        for candidate, mapped_key in _FIELD_KEY_MAP.items():
                            if _normalize_key(candidate) == normalized_key:
                                column_map[mapped_key] = col_index
                                break
                    continue

            if header_row_values is None:
                continue

            if not column_map:
                for candidate, mapped_key in _FIELD_KEY_MAP.items():
                    normalized_candidate = _normalize_key(candidate)
                    for col_index, cell_value in enumerate(header_row_values):
                        if _normalize_key(cell_value) == normalized_candidate:
                            column_map[mapped_key] = col_index
                            break
                if not column_map:
                    for index, header in enumerate(headers):
                        mapped_key = _FIELD_KEY_MAP.get(header)
                        if mapped_key:
                            column_map[mapped_key] = index

            if looks_like_header_row(row_values, headers):
                continue

            row_data: Dict[str, str] = dict(_DEFAULT_ROW)
            has_value = False
            for key, header in _COLUMN_HEADERS.items():
                column_index = column_map.get(key)
                if column_index is None:
                    continue
                value = (
                    row_values[column_index]
                    if column_index < len(row_values)
                    else None
                )
                text = _normalize_value(value)
                if text:
                    has_value = True
                row_data[key] = text

            if has_value:
                extracted_rows.append(row_data)

        if header_row_index is not None:
            start_row = header_row_index + 1
    finally:
        workbook.close()

    if not sheet_title:
        sheet_title = "결함리포트"

    normalized_rows = normalize_defect_report_rows(extracted_rows)
    return sheet_title, start_row, list(headers), normalized_rows


def build_defect_report_rows_csv(rows: Sequence[Mapping[str, Any]]) -> str:
    normalized_rows = normalize_defect_report_rows(rows)

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=list(DEFECT_REPORT_EXPECTED_HEADERS),
        lineterminator="\n",
        delimiter=AI_CSV_DELIMITER,
    )
    writer.writeheader()

    for row in normalized_rows:
        entry = {header: "" for header in DEFECT_REPORT_EXPECTED_HEADERS}
        for key, header in _COLUMN_HEADERS.items():
            entry[header] = row.get(key, "")
        writer.writerow(entry)

    return output.getvalue()


def prepare_defect_report_response(
    *,
    file_id: str,
    file_name: str,
    sheet_name: str,
    start_row: int,
    headers: Sequence[str],
    rows: Sequence[Mapping[str, Any]],
    modified_time: Optional[str],
) -> Dict[str, Any]:
    return {
        "fileId": file_id,
        "fileName": file_name,
        "sheetName": sheet_name,
        "startRow": start_row,
        "headers": list(headers),
        "rows": [dict(row) for row in rows],
        "modifiedTime": modified_time,
    }


def serialize_defect_report_images(images: Iterable[DefectReportImage]) -> List[Dict[str, Any]]:
    serialized: List[Dict[str, Any]] = []
    for image in images:
        serialized.append(
            {
                "fileName": image.file_name,
                "contentType": image.content_type,
                "size": len(image.content),
            }
        )
    return serialized

