"""Workflow helpers for Drive testcase spreadsheets."""
from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Optional, Sequence, Tuple

from fastapi import HTTPException

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]

from ..excel_templates.models import TESTCASE_EXPECTED_HEADERS, TESTCASE_START_ROW
from ..excel_templates.utils import AI_CSV_DELIMITER
from .naming import drive_name_matches, looks_like_header_row, normalize_drive_text, squash_drive_text

__all__ = [
    "parse_testcase_workbook",
    "build_testcase_rows_csv",
    "prepare_testcase_response",
]

_TESTCASE_SHEET_CANDIDATES: Tuple[str, ...] = (
    "테스트케이스",
    "테스트 케이스",
    "testcase",
    "test cases",
)

_HEADER_KEY_MAP = {
    "대분류": "majorCategory",
    "중분류": "middleCategory",
    "소분류": "minorCategory",
    "테스트 케이스 ID": "testcaseId",
    "TC_ID": "testcaseId",
    "테스트케이스 ID": "testcaseId",
    "테스트 시나리오": "scenario",
    "테스트 시나리오(상세)": "scenario",
    "입력(사전조건 포함)": "input",
    "기대 출력(사후조건 포함)": "expected",
    "테스트 결과": "result",
    "상세 테스트 결과": "detail",
    "비고": "note",
}

_DEFAULT_ROW: Dict[str, str] = {
    "majorCategory": "",
    "middleCategory": "",
    "minorCategory": "",
    "testcaseId": "",
    "scenario": "",
    "input": "",
    "expected": "",
    "result": "",
    "detail": "",
    "note": "",
}


def _count_header_matches(values: Sequence[Any], expected: Sequence[str]) -> int:
    if not values:
        return 0

    normalized_values = [
        normalize_drive_text(str(value)) if value is not None else ""
        for value in values
    ]
    squashed_values = [squash_drive_text(value) for value in normalized_values]
    normalized_expected = [normalize_drive_text(name) for name in expected]
    squashed_expected = [squash_drive_text(name) for name in normalized_expected]

    matches = 0
    for expected_value, expected_squashed in zip(normalized_expected, squashed_expected):
        if not expected_value and not expected_squashed:
            continue

        for actual_value, actual_squashed in zip(normalized_values, squashed_values):
            if not actual_value and not actual_squashed:
                continue

            normalized_match = (
                bool(expected_value)
                and bool(actual_value)
                and (
                    actual_value == expected_value
                    or expected_value in actual_value
                    or actual_value in expected_value
                )
            )
            squashed_match = (
                bool(expected_squashed)
                and bool(actual_squashed)
                and expected_squashed in actual_squashed
            )

            if normalized_match or squashed_match:
                matches += 1
                break

    return matches


def _normalize_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _resolve_column_index(header_row: Sequence[Any], expected_header: str) -> Optional[int]:
    if not header_row:
        return None

    for index, value in enumerate(header_row):
        if value is None:
            continue
        text = str(value)
        if drive_name_matches(text, expected_header):
            return index

    try:
        return header_row.index(expected_header)
    except ValueError:
        return None


def parse_testcase_workbook(workbook_bytes: bytes) -> Tuple[str, int, List[str], List[Dict[str, str]]]:
    """Extract testcase rows from the provided workbook bytes."""

    if load_workbook is None:  # pragma: no cover
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 필요합니다.")

    buffer = io.BytesIO(workbook_bytes)
    try:
        workbook = load_workbook(buffer, data_only=True)
    except Exception as exc:  # pragma: no cover - safety net
        raise HTTPException(status_code=500, detail="엑셀 파일을 읽는 중 오류가 발생했습니다.") from exc

    headers = list(TESTCASE_EXPECTED_HEADERS)
    extracted_rows: List[Dict[str, str]] = []
    sheet_title = ""
    start_row = 1
    column_map: Dict[str, int] = {}

    try:
        sheet = workbook.active
        selected_title = sheet.title
        for candidate in _TESTCASE_SHEET_CANDIDATES:
            matched = False
            for title in workbook.sheetnames:
                if drive_name_matches(title, candidate):
                    try:
                        sheet = workbook[title]
                        selected_title = sheet.title
                        matched = True
                        break
                    except KeyError:
                        continue
            if matched:
                break

        sheet_title = selected_title or ""
        max_col = max(len(headers), sheet.max_column or len(headers))
        rows_snapshot: List[Tuple[int, Sequence[Any]]] = [
            (idx, row if isinstance(row, Sequence) else tuple())
            for idx, row in enumerate(
                sheet.iter_rows(min_row=1, max_col=max_col, values_only=True),
                start=1,
            )
        ]

        strong_threshold = max(1, len(headers) - 1)
        fallback_threshold = max(1, len(headers) // 2)
        header_row_values: Optional[Sequence[Any]] = None
        header_row_index: Optional[int] = None
        best_candidate: Optional[Tuple[int, Sequence[Any], int]] = None

        for idx, row_values in rows_snapshot:
            if not any(value is not None for value in row_values):
                continue

            normalized = [_normalize_value(value) for value in row_values]
            match_count = _count_header_matches(normalized, headers)

            if match_count >= strong_threshold:
                header_row_values = normalized
                header_row_index = idx
                break

            if match_count > 0:
                if best_candidate is None or match_count > best_candidate[2]:
                    best_candidate = (idx, normalized, match_count)

        if header_row_values is None and best_candidate and best_candidate[2] >= fallback_threshold:
            header_row_index, header_row_values, _ = best_candidate

        if header_row_values is None:
            header_row_values = list(headers)
            header_row_index = TESTCASE_START_ROW - 1

        first_data_row_index: Optional[int] = None

        for header_name in headers:
            column_index = _resolve_column_index(header_row_values, header_name)
            if column_index is None:
                column_index = headers.index(header_name)
            column_map[header_name] = column_index

        for idx, row_values in rows_snapshot:
            if idx <= (header_row_index or 0):
                continue

            if not any(value is not None for value in row_values):
                continue

            normalized = [_normalize_value(value) for value in row_values]
            if looks_like_header_row(normalized, headers):
                continue

            if first_data_row_index is None:
                first_data_row_index = idx

            row_data = dict(_DEFAULT_ROW)
            has_values = False

            for header_name in headers:
                column_index = column_map.get(header_name)
                if column_index is None:
                    continue
                cell_value = (
                    row_values[column_index]
                    if column_index < len(row_values)
                    else None
                )
                text = _normalize_value(cell_value)
                if text:
                    has_values = True

                key = _HEADER_KEY_MAP.get(header_name)
                if key:
                    row_data[key] = text

            if not has_values:
                continue

            extracted_rows.append(row_data)

        if header_row_index is not None:
            start_row = header_row_index + 1
        if first_data_row_index is not None:
            start_row = first_data_row_index
    finally:
        workbook.close()

    if not sheet_title:
        sheet_title = "테스트케이스"

    return sheet_title, start_row, headers, extracted_rows


def build_testcase_rows_csv(rows: Sequence[Dict[str, str]]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=list(TESTCASE_EXPECTED_HEADERS),
        lineterminator="\n",
        delimiter=AI_CSV_DELIMITER,
    )
    writer.writeheader()

    for row in rows:
        major = _normalize_value(row.get("majorCategory"))
        middle = _normalize_value(row.get("middleCategory"))
        minor = _normalize_value(row.get("minorCategory"))
        testcase_id = _normalize_value(row.get("testcaseId"))
        scenario = _normalize_value(row.get("scenario"))
        input_value = _normalize_value(row.get("input"))
        expected = _normalize_value(row.get("expected"))
        result = _normalize_value(row.get("result"))
        detail = _normalize_value(row.get("detail"))
        note = _normalize_value(row.get("note"))

        if not any([major, middle, minor, testcase_id, scenario, input_value, expected, result, detail, note]):
            continue

        writer.writerow(
            {
                "대분류": major,
                "중분류": middle,
                "소분류": minor,
                "테스트 케이스 ID": testcase_id,
                "테스트 시나리오": scenario,
                "입력(사전조건 포함)": input_value,
                "기대 출력(사후조건 포함)": expected,
                "테스트 결과": result,
                "상세 테스트 결과": detail,
                "비고": note,
            }
        )

    return output.getvalue()


def prepare_testcase_response(
    *,
    file_id: str,
    file_name: str,
    sheet_name: str,
    start_row: int,
    headers: Sequence[str],
    rows: Sequence[Dict[str, str]],
    modified_time: Optional[str],
) -> Dict[str, Any]:
    response: Dict[str, Any] = {
        "fileId": file_id,
        "fileName": file_name,
        "sheetName": sheet_name,
        "startRow": start_row,
        "headers": list(headers),
        "rows": list(rows),
        "modifiedTime": modified_time,
    }
    return response
