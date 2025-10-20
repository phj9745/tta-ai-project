from __future__ import annotations

import csv
import io
import logging
import math
import posixpath
import re
import zipfile
from collections import defaultdict
from copy import deepcopy
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from xml.etree import ElementTree as ET

from fastapi import HTTPException, UploadFile

from openpyxl import load_workbook
from openpyxl.chart.reference import Reference
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.worksheet import Worksheet

from .google_drive import GoogleDriveService, XLSX_MIME_TYPE

logger = logging.getLogger(__name__)

TEMPLATE_PATH = (
    Path(__file__).resolve().parents[2]
    / "template"
    / "다.수행"
    / "성능시험"
    / "GS-X-XX-XXXX 성능시험.xlsx"
)


class PerformanceReportError(RuntimeError):
    """Raised when the performance report cannot be assembled."""


class UnsupportedFormatError(PerformanceReportError):
    """Raised when a rawdata file cannot be parsed."""


class OSType(Enum):
    WINDOWS = "Windows"
    LINUX = "Linux"
    ANDROID = "Android"
    IOS = "iOS"

    @property
    def sheet_name(self) -> str:
        return self.value

    @property
    def display_name(self) -> str:
        return self.value


WINDOWS_PATTERN = re.compile(r"\\Processor\(_Total\)\\% Processor Time", re.IGNORECASE)
LINUX_PATTERN = re.compile(r"procs\s+-+memory-+\s+-+swap--", re.IGNORECASE)
ANDROID_HINT_PATTERN = re.compile(r"\bandroid\b|\btop\b|\bdumpsys\b", re.IGNORECASE)
IOS_HINT_PATTERN = re.compile(r"Total\s+Load%|Memory\s+Used", re.IGNORECASE)

TIMESTAMP_COLUMN_KEYS: Tuple[str, ...] = ("timestamp", "time", "datetime")


@dataclass(slots=True)
class RawDataset:
    filename: str
    os_type: OSType
    records: List[Dict[str, Any]]


@dataclass(slots=True)
class GeneratedWorkbook:
    filename: str
    content: bytes
    uploaded_file_id: Optional[str] = None


def detect_os(filename: str, text: str) -> OSType:
    """
    Detect the operating system for a rawdata file.
    """
    if WINDOWS_PATTERN.search(text):
        return OSType.WINDOWS
    if LINUX_PATTERN.search(text):
        return OSType.LINUX

    lower_name = filename.lower()
    if "android" in lower_name or ANDROID_HINT_PATTERN.search(text):
        return OSType.ANDROID
    if "ios" in lower_name or "iphone" in lower_name or IOS_HINT_PATTERN.search(text):
        return OSType.IOS

    raise UnsupportedFormatError(
        f"운영체제를 자동으로 판별할 수 없습니다: {filename}. "
        "데이터에 OS 식별 문자열을 포함시키거나 파일명에 운영체제를 명시해 주세요."
    )


def _read_text_with_fallback(raw_bytes: bytes) -> str:
    """
    Decode bytes from uploads. Try utf-8 first, fallback to cp949 and latin-1.
    """
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8", errors="ignore")


def parse_windows_perfmon(text: str) -> List[Dict[str, Any]]:
    reader = csv.reader(io.StringIO(text.strip()))
    rows = list(reader)
    if not rows or len(rows) < 2:
        raise UnsupportedFormatError("Perfmon CSV 데이터에서 레코드를 찾을 수 없습니다.")

    header = rows[0]
    column_map: Dict[str, int] = {col.strip(): idx for idx, col in enumerate(header)}

    timestamp_idx = 0
    cpu_idx: Optional[int] = None
    private_idx: Optional[int] = None

    for name, idx in column_map.items():
        key = name.lower()
        if cpu_idx is None and "% processor time" in key:
            cpu_idx = idx
        if private_idx is None and "private bytes" in key:
            private_idx = idx

    if cpu_idx is None or private_idx is None:
        raise UnsupportedFormatError("Perfmon CSV에서 '% Processor Time' 혹은 'Private Bytes' 열을 찾을 수 없습니다.")

    datasets: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if len(row) <= max(timestamp_idx, cpu_idx, private_idx):
            continue
        timestamp = row[timestamp_idx].strip()
        private_raw = row[private_idx].strip()
        cpu_raw = row[cpu_idx].strip()
        if not timestamp:
            continue

        record: Dict[str, Any] = {
            "timestamp": timestamp,
            "private_bytes": _to_number(private_raw),
            "cpu_percent": _to_number(cpu_raw),
        }
        datasets.append(record)

    if not datasets:
        raise UnsupportedFormatError("Perfmon CSV에서 유효한 성능 데이터를 찾지 못했습니다.")
    return datasets


VMSTAT_COLUMNS = (
    "r",
    "b",
    "swpd",
    "free",
    "buff",
    "cache",
    "si",
    "so",
    "bi",
    "bo",
    "in",
    "cs",
    "us",
    "sy",
    "id",
    "wa",
    "st",
)


def parse_vmstat(text: str) -> List[Dict[str, Any]]:
    lines = [line.rstrip("\n") for line in text.splitlines()]
    if not lines:
        raise UnsupportedFormatError("vmstat 결과가 비어 있습니다.")

    header_index = None
    for idx, line in enumerate(lines):
        if LINUX_PATTERN.search(line):
            header_index = idx
            break

    if header_index is None or header_index + 1 >= len(lines):
        raise UnsupportedFormatError("vmstat 헤더를 찾지 못했습니다.")

    data_start = header_index + 2
    records: List[Dict[str, Any]] = []
    for raw_line in lines[data_start:]:
        stripped = raw_line.strip()
        if not stripped:
            continue
        parts = stripped.split()
        if len(parts) < len(VMSTAT_COLUMNS):
            continue

        stats = parts[: len(VMSTAT_COLUMNS)]
        timestamp = " ".join(parts[len(VMSTAT_COLUMNS):]).strip()
        record: Dict[str, Any] = {}
        for column, value in zip(VMSTAT_COLUMNS, stats):
            record[column] = _to_number(value)
        if timestamp:
            record["timestamp"] = timestamp
        records.append(record)

    if not records:
        raise UnsupportedFormatError("vmstat 데이터에서 유효한 측정값을 찾지 못했습니다.")
    return records


def parse_android(text: str) -> List[Dict[str, Any]]:
    """Android 성능 데이터는 Linux(vmstat)와 구조가 동일하다고 가정한다."""
    return parse_vmstat(text)


def parse_ios(text: str) -> List[Dict[str, Any]]:
    reader = csv.reader(io.StringIO(text.strip()))
    rows = list(reader)
    if len(rows) < 2:
        raise UnsupportedFormatError("iOS 성능 CSV에서 데이터를 찾지 못했습니다.")

    header = [cell.strip() for cell in rows[0]]
    column_map = {name.lower(): idx for idx, name in enumerate(header)}
    memory_idx = None
    load_idx = None
    timestamp_idx: Optional[int] = None
    for name, idx in column_map.items():
        if "memory" in name and "used" in name and memory_idx is None:
            memory_idx = idx
        if "load" in name and "%" in name and load_idx is None:
            load_idx = idx
        if any(key in name for key in TIMESTAMP_COLUMN_KEYS) and timestamp_idx is None:
            timestamp_idx = idx

    if memory_idx is None or load_idx is None:
        raise UnsupportedFormatError("iOS CSV에서 'Memory Used' 혹은 'Total Load%' 열을 찾지 못했습니다.")

    datasets: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if len(row) <= max(memory_idx, load_idx):
            continue
        record: Dict[str, Any] = {
            "memory_used": _to_number(row[memory_idx].strip()),
            "total_load": _to_number(row[load_idx].strip()),
        }
        if timestamp_idx is not None and timestamp_idx < len(row):
            record["timestamp"] = row[timestamp_idx].strip()
        datasets.append(record)

    if not datasets:
        raise UnsupportedFormatError("iOS CSV에서 유효한 데이터를 찾지 못했습니다.")
    return datasets


def parse_raw_dataset(filename: str, raw_bytes: bytes, forced_os: Optional[OSType] = None) -> RawDataset:
    text = _read_text_with_fallback(raw_bytes)
    if forced_os is not None:
        os_type = forced_os
    else:
        os_type = detect_os(filename, text)

    if os_type is OSType.WINDOWS:
        records = parse_windows_perfmon(text)
    elif os_type is OSType.LINUX:
        records = parse_vmstat(text)
    elif os_type is OSType.ANDROID:
        records = parse_android(text)
    elif os_type is OSType.IOS:
        records = parse_ios(text)
    else:  # pragma: no cover - defensive
        raise UnsupportedFormatError(f"지원하지 않는 운영체제입니다: {os_type}")

    logger.info("Rawdata parsed", extra={"file": filename, "os": os_type.value, "records": len(records)})
    return RawDataset(filename=filename, os_type=os_type, records=records)


def _to_number(value: str) -> Optional[float]:
    if not value:
        return None
    stripped = value.strip()
    if not stripped:
        return None
    try:
        number = float(stripped.replace(",", ""))
    except ValueError:
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


class PerformanceReportBuilder:
    START_ROW = 4

    def __init__(self, template_bytes: bytes) -> None:
        self._template_bytes = template_bytes
        self._workbook = load_workbook(io.BytesIO(template_bytes))
        self._remove_external_links()
        self._chart_reference_jobs: List[Tuple[str, str]] = []

    @property
    def workbook(self):
        return self._workbook

    def build(self, datasets: Sequence[RawDataset]) -> bytes:
        if not datasets:
            raise PerformanceReportError("최소 한 개의 Rawdata 파일이 필요합니다.")

        grouped: Dict[OSType, List[RawDataset]] = defaultdict(list)
        for dataset in datasets:
            grouped[dataset.os_type].append(dataset)

        self._prepare_sheets(grouped)

        for os_type, entries in grouped.items():
            template_sheet = self._workbook[os_type.sheet_name]
            sheet_clones = self._clone_os_sheets(
                template_sheet,
                count=len(entries),
                base_title=os_type.sheet_name,
            )
            for index, (sheet, dataset) in enumerate(zip(sheet_clones, entries), start=1):
                original_title = sheet.title
                sheet.title = f"{os_type.display_name} #{index}"
                if original_title != sheet.title:
                    self._rename_defined_names(original_title, sheet.title)
                self._chart_reference_jobs.append((sheet.title, os_type.sheet_name))
                logger.info(
                    "Filling sheet",
                    extra={"sheet": sheet.title, "records": len(dataset.records), "file": dataset.filename},
                )
                self._insert_records(sheet, dataset)
                self._update_charts(sheet, len(dataset.records))

        self._cleanup_unused_os_sheets(grouped)

        buffer = io.BytesIO()
        self._workbook.save(buffer)
        buffer.seek(0)
        raw_bytes = buffer.getvalue()

        # 차트 XML 전역 재작성(원본 템플릿 시트 참조 제거)
        rename_targets = {
            (new_title, base_title)
            for new_title, base_title in self._chart_reference_jobs
            if new_title != base_title
        }
        updated_bytes = _rewrite_chart_references(raw_bytes, rename_targets)
        logger.info("Performance workbook assembled", extra={"sheets": self._workbook.sheetnames})
        return updated_bytes

    def _prepare_sheets(self, grouped: Dict[OSType, List[RawDataset]]) -> None:
        for os_type, datasets in grouped.items():
            if os_type.sheet_name not in self._workbook.sheetnames:
                raise PerformanceReportError(f"템플릿에서 '{os_type.sheet_name}' 시트를 찾지 못했습니다.")
            if not datasets:
                continue
            template_sheet = self._workbook[os_type.sheet_name]
            self._ensure_chart_templates(template_sheet)

    def _clone_os_sheets(self, template_sheet: Worksheet, count: int, *, base_title: str) -> List[Worksheet]:
        sheets: List[Worksheet] = []
        if count <= 0:
            return sheets
        sheets.append(template_sheet)
        template_chart_snapshot = list(template_sheet._charts or [])
        for _ in range(count - 1):
            clone = self._workbook.copy_worksheet(template_sheet)
            if template_chart_snapshot:
                clone_charts(template_sheet, clone)
            sheets.append(clone)
        return sheets

    def _ensure_chart_templates(self, sheet: Worksheet) -> None:
        if not getattr(sheet, "_charts", None):
            return
        if sheet._charts:  # pragma: no cover - template already prepared
            return
        raise PerformanceReportError(f"'{sheet.title}' 시트에서 차트 템플릿을 찾지 못했습니다.")

    def _insert_records(self, sheet: Worksheet, dataset: RawDataset) -> None:
        if dataset.os_type is OSType.WINDOWS:
            self._insert_windows_records(sheet, dataset.records)
        elif dataset.os_type is OSType.LINUX:
            self._insert_linux_records(sheet, dataset.records)
        elif dataset.os_type is OSType.ANDROID:
            self._insert_android_records(sheet, dataset.records)
        elif dataset.os_type is OSType.IOS:
            self._insert_ios_records(sheet, dataset.records)
        else:  # pragma: no cover - defensive
            raise PerformanceReportError(f"지원되지 않는 시트 유형입니다: {dataset.os_type}")

    def _insert_windows_records(self, sheet: Worksheet, records: Sequence[Dict[str, Any]]) -> None:
        row = self.START_ROW
        for record in records:
            sheet.cell(row=row, column=4).value = record.get("timestamp")
            sheet.cell(row=row, column=5).value = record.get("cpu_percent")
            sheet.cell(row=row, column=6).value = record.get("private_bytes")
            row += 1
        self._clear_tail(sheet, start_row=row, columns=(4, 5, 6))

    def _insert_linux_records(self, sheet: Worksheet, records: Sequence[Dict[str, Any]]) -> None:
        row = self.START_ROW
        for record in records:
            sheet.cell(row=row, column=4).value = record.get("us")
            sheet.cell(row=row, column=5).value = record.get("sy")
            sheet.cell(row=row, column=6).value = record.get("free")
            sheet.cell(row=row, column=7).value = record.get("buff")
            sheet.cell(row=row, column=8).value = record.get("cache")
            sheet.cell(row=row, column=9).value = record.get("bi")
            sheet.cell(row=row, column=10).value = record.get("bo")
            row += 1
        self._clear_tail(sheet, start_row=row, columns=(4, 5, 6, 7, 8, 9, 10))

    def _insert_android_records(self, sheet: Worksheet, records: Sequence[Dict[str, Any]]) -> None:
        row = self.START_ROW
        for record in records:
            sheet.cell(row=row, column=4).value = record.get("us")
            sheet.cell(row=row, column=5).value = record.get("sy")
            sheet.cell(row=row, column=6).value = record.get("free")
            sheet.cell(row=row, column=7).value = record.get("buff")
            sheet.cell(row=row, column=8).value = record.get("cache")
            sheet.cell(row=row, column=18).value = record.get("timestamp")
            row += 1
        self._clear_tail(sheet, start_row=row, columns=(4, 5, 6, 7, 8, 18))

    def _insert_ios_records(self, sheet: Worksheet, records: Sequence[Dict[str, Any]]) -> None:
        row = self.START_ROW
        for record in records:
            sheet.cell(row=row, column=4).value = record.get("memory_used")
            sheet.cell(row=row, column=5).value = record.get("total_load")
            sheet.cell(row=row, column=16).value = record.get("timestamp")
            row += 1
        self._clear_tail(sheet, start_row=row, columns=(4, 5, 16))

    def _clear_tail(self, sheet: Worksheet, start_row: int, columns: Sequence[int]) -> None:
        max_row = sheet.max_row
        if start_row > max_row:
            return
        for row in range(start_row, max_row + 1):
            for column in columns:
                sheet.cell(row=row, column=column).value = None

    def _update_charts(self, sheet: Worksheet, record_count: int) -> None:
        if record_count <= 0:
            logger.warning("Skipping chart update for empty sheet", extra={"sheet": sheet.title})
            return

        end_row = self.START_ROW + record_count - 1
        charts = list(sheet._charts or [])
        if not charts:
            logger.warning("No charts attached to sheet", extra={"sheet": sheet.title})
            return

        for chart in charts:
            for series in list(getattr(chart, "series", [])):
                # Values 범위가 Reference면 행 범위 늘리기
                if hasattr(series, "values"):
                    if isinstance(series.values, Reference):
                        series.values = Reference(
                            sheet,
                            min_col=series.values.min_col,
                            min_row=series.values.min_row,
                            max_col=series.values.max_col,
                            max_row=end_row,
                        )
                    elif isinstance(series.values, str):
                        # 문자열 수식인 경우 시트명만 보정
                        series.values = self._rewrite_sheet_reference(series.values, sheet.title)

                # XValues 범위가 Reference면 행 범위 늘리기
                if hasattr(series, "xvalues"):
                    if isinstance(series.xvalues, Reference):
                        series.xvalues = Reference(
                            sheet,
                            min_col=series.xvalues.min_col,
                            min_row=series.xvalues.min_row,
                            max_col=series.xvalues.max_col,
                            max_row=end_row,
                        )
                    elif isinstance(series.xvalues, str):
                        series.xvalues = self._rewrite_sheet_reference(series.xvalues, sheet.title)

                # 시리즈 라벨 수식의 시트명도 보정
                label = getattr(series, "tx", None)
                if label is not None:
                    str_ref = getattr(label, "strRef", None)
                    if str_ref is not None and getattr(str_ref, "f", None):
                        str_ref.f = self._rewrite_sheet_reference(str_ref.f, sheet.title)

    def _cleanup_unused_os_sheets(self, grouped: Dict[OSType, List[RawDataset]]) -> None:
        expected_titles = {"시나리오", "응답시간"}
        for os_type, datasets in grouped.items():
            for idx in range(1, len(datasets) + 1):
                expected_titles.add(f"{os_type.display_name} #{idx}")

        for sheet_name in list(self._workbook.sheetnames):
            if sheet_name in expected_titles:
                continue
            if sheet_name in (ost.sheet_name for ost in OSType):
                del self._workbook[sheet_name]

    def _remove_external_links(self) -> None:
        if hasattr(self._workbook, "_external_links") and self._workbook._external_links:
            logger.debug("Removing %d external links from workbook", len(self._workbook._external_links))
            self._workbook._external_links = []

    def _rename_defined_names(self, old_title: str, new_title: str) -> None:
        defined_names = getattr(self._workbook, "defined_names", None)
        if not defined_names:
            return

        quoted_old = quote_sheetname(old_title)
        quoted_new = quote_sheetname(new_title)

        for definition in list(getattr(defined_names, "definedName", [])):
            txt = definition.text or ""
            # 시트명 참조가 문자열 어디에 있어도 전역 치환
            replaced = txt.replace(f"{old_title}!", f"{quoted_new}!") \
                          .replace(f"{quoted_old}!", f"{quoted_new}!")
            if replaced != txt:
                definition.text = replaced

    @staticmethod
    def _rewrite_sheet_reference(formula: str, sheet_title: str) -> str:
        if "!" not in formula:
            return formula
        quoted_new = quote_sheetname(sheet_title)
        if formula.startswith(f"{sheet_title}!") or formula.startswith(f"{quoted_new}!"):
            return formula
        _, rest = formula.split("!", 1)
        return f"{quoted_new}!{rest}"


def clone_charts(source: Worksheet, target: Worksheet) -> None:
    target._charts = []
    for chart in source._charts:
        cloned_chart = deepcopy(chart)
        target.add_chart(cloned_chart, chart.anchor)


class PerformanceReportService:
    """Facade that orchestrates Drive integration and workbook construction."""

    def __init__(self, drive_service: GoogleDriveService) -> None:
        self._drive_service = drive_service

    async def generate_report(
        self,
        *,
        project_id: str,
        uploads: Sequence[UploadFile],
        google_id: Optional[str],
        os_hints: Optional[Sequence[Optional[str]]] = None,
    ) -> GeneratedWorkbook:
        if not uploads:
            raise HTTPException(status_code=422, detail="최소 한 개의 Rawdata 파일을 업로드해 주세요.")

        datasets: List[RawDataset] = []
        ambiguous: List[Dict[str, Any]] = []
        hints = list(os_hints or [])

        for index, upload in enumerate(uploads):
            filename = upload.filename or f"rawdata-{index + 1}"
            try:
                raw_bytes = await upload.read()
            finally:
                await upload.close()
            if not raw_bytes:
                raise HTTPException(status_code=422, detail=f"파일이 비어 있습니다: {filename}")
            forced_os = None
            if index < len(hints):
                forced_os = self._parse_os_hint(hints[index])
            try:
                dataset = parse_raw_dataset(filename, raw_bytes, forced_os=forced_os)
            except UnsupportedFormatError as exc:
                if forced_os is not None:
                    raise HTTPException(
                        status_code=422,
                        detail=f"{filename} 파일을 {forced_os.value} 형식으로 읽을 수 없습니다.",
                    ) from exc
                ambiguous.append({"filename": filename, "index": index})
                continue
            datasets.append(dataset)

        if ambiguous:
            raise HTTPException(
                status_code=422,
                detail={
                    "code": "os_selection_required",
                    "message": "자동으로 운영체제를 판별할 수 없는 Rawdata 파일이 있습니다. 운영체제를 선택해 주세요.",
                    "files": ambiguous,
                },
            )

        try:
            template_bytes = TEMPLATE_PATH.read_bytes()
        except FileNotFoundError as exc:
            logger.exception("Performance template file missing.", extra={"path": str(TEMPLATE_PATH)})
            raise HTTPException(status_code=500, detail="성능시험 템플릿 파일을 찾지 못했습니다. 관리자에게 문의해주세요.") from exc

        builder = PerformanceReportBuilder(template_bytes)
        try:
            workbook_bytes = builder.build(datasets)
        except PerformanceReportError as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc

        project_number = await self._drive_service.get_project_exam_number(
            project_id=project_id,
            google_id=google_id,
        )
        filename = f"{project_number} 성능시험.xlsx"

        target_folder_id = await self._drive_service.ensure_project_subfolder(
            project_id=project_id,
            path=("다.수행", "성능시험"),
            google_id=google_id,
        )

        # 빈 템플릿 정리
        await self._drive_service.delete_files_by_name(
            parent_id=target_folder_id,
            name="GS-X-XX-XXXX 성능시험.xlsx",
            google_id=google_id,
        )

        upload_result = await self._drive_service.upload_file_to_folder(
            parent_id=target_folder_id,
            file_name=filename,
            content=workbook_bytes,
            content_type=XLSX_MIME_TYPE,
            google_id=google_id,
        )
        file_id = upload_result.get("id")

        logger.info(
            "Performance workbook uploaded",
            extra={"project_id": project_id, "file_id": file_id, "filename": filename},
        )

        return GeneratedWorkbook(filename=filename, content=workbook_bytes, uploaded_file_id=file_id)

    @staticmethod
    def _parse_os_hint(value: Optional[str]) -> Optional[OSType]:
        if value is None:
            return None
        normalized = value.strip()
        if not normalized:
            return None
        lowered = normalized.lower()
        for candidate in OSType:
            if candidate.value.lower() == lowered:
                return candidate
        raise HTTPException(status_code=422, detail=f"지원하지 않는 운영체제 지정입니다: {normalized}")


# -------------------------
# Chart XML rewrite helpers
# -------------------------

def _rewrite_chart_references(workbook_bytes: bytes, rename_targets: Iterable[Tuple[str, str]]) -> bytes:
    targets = [target for target in rename_targets if target[0] != target[1]]
    if not targets:
        return workbook_bytes

    source_buffer = io.BytesIO(workbook_bytes)
    output_buffer = io.BytesIO()

    with zipfile.ZipFile(source_buffer, "r") as zin:
        if "xl/workbook.xml" not in zin.namelist() or "xl/_rels/workbook.xml.rels" not in zin.namelist():
            return workbook_bytes

        workbook_tree = ET.fromstring(zin.read("xl/workbook.xml"))
        rel_tree = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        rel_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

        id_to_target: Dict[str, str] = {}
        for rel in rel_tree.findall("rel:Relationship", rel_ns):
            r_id = rel.get("Id")
            target = rel.get("Target")
            if r_id and target:
                id_to_target[r_id] = target

        sheet_to_path: Dict[str, str] = {}
        sheets_elem = workbook_tree.find("s:sheets", ns)
        if sheets_elem is None:
            return workbook_bytes
        for sheet in sheets_elem.findall("s:sheet", ns):
            name = sheet.get("name")
            r_id = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if name and r_id and r_id in id_to_target:
                sheet_to_path[name] = id_to_target[r_id]

        chart_overrides: Dict[str, bytes] = {}

        for new_title, base_title in targets:
            sheet_rel_path = sheet_to_path.get(new_title)
            if not sheet_rel_path:
                continue
            sheet_xml_path = _normalize_zip_path(f"xl/{sheet_rel_path}")
            rel_path = _sheet_relationship_path(sheet_xml_path)
            if rel_path not in zin.namelist():
                continue

            # NOTE: Target of a .rels is resolved RELATIVE TO THE .rels FILE'S FOLDER
            sheet_rel_tree = ET.fromstring(zin.read(rel_path))
            drawing_targets: List[str] = []
            for rel in sheet_rel_tree.findall("rel:Relationship", rel_ns):
                rel_type = rel.get("Type")
                if rel_type == "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing":
                    target = rel.get("Target")
                    if target:
                        drawing_targets.append(
                            _normalize_zip_path(posixpath.join(posixpath.dirname(rel_path), target))
                        )

            for drawing_path in drawing_targets:
                if drawing_path not in zin.namelist():
                    continue
                drawing_tree = ET.fromstring(zin.read(drawing_path))
                drawing_rel_path = _drawing_relationship_path(drawing_path)
                if drawing_rel_path not in zin.namelist():
                    continue

                drawing_rel_tree = ET.fromstring(zin.read(drawing_rel_path))
                chart_id_to_target: Dict[str, str] = {}
                for rel in drawing_rel_tree.findall("rel:Relationship", rel_ns):
                    rel_type = rel.get("Type")
                    if rel_type == "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart":
                        r_id = rel.get("Id")
                        target = rel.get("Target")
                        if r_id and target:
                            chart_id_to_target[r_id] = target

                chart_ns = {
                    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
                    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
                }
                # Cover all possible anchor types
                for anchor_tag in ("xdr:twoCellAnchor", "xdr:oneCellAnchor", "xdr:absoluteAnchor"):
                    for anchor in drawing_tree.findall(anchor_tag, chart_ns):
                        chart_elem = anchor.find(".//c:chart", chart_ns)
                        if chart_elem is None:
                            continue
                        r_id = chart_elem.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                        if not r_id or r_id not in chart_id_to_target:
                            continue

                        # NOTE: Chart Target is relative to the DRAWING .rels folder
                        chart_target = _normalize_zip_path(
                            posixpath.join(posixpath.dirname(drawing_rel_path), chart_id_to_target[r_id])
                        )
                        if chart_target not in zin.namelist():
                            continue

                        original_xml = chart_overrides.get(chart_target)
                        if original_xml is None:
                            original_xml = zin.read(chart_target)
                        updated_xml = _replace_sheet_references_in_xml(original_xml, base_title, new_title)
                        chart_overrides[chart_target] = updated_xml

        if not chart_overrides:
            return workbook_bytes

        with zipfile.ZipFile(output_buffer, "w") as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename in chart_overrides:
                    data = chart_overrides[info.filename]
                zout.writestr(info, data)

    return output_buffer.getvalue()


def _sheet_relationship_path(sheet_xml_path: str) -> str:
    directory, filename = posixpath.split(sheet_xml_path)
    rel_directory = posixpath.join(directory, "_rels")
    return posixpath.join(rel_directory, f"{filename}.rels")


def _drawing_relationship_path(drawing_xml_path: str) -> str:
    directory, filename = posixpath.split(drawing_xml_path)
    rel_directory = posixpath.join(directory, "_rels")
    return posixpath.join(rel_directory, f"{filename}.rels")


def _normalize_zip_path(path: str) -> str:
    parts: List[str] = []
    for segment in path.replace("\\", "/").split("/"):
        if segment in ("", "."):
            continue
        if segment == "..":
            if parts:
                parts.pop()
            continue
        parts.append(segment)
    return "/".join(parts)


def _replace_sheet_references_in_xml(content: bytes, base_title: str, new_title: str) -> bytes:
    text = content.decode("utf-8")
    quoted_new = quote_sheetname(new_title)
    replacements = [
        (f"{base_title}!", f"{quoted_new}!"),
        (f"{quote_sheetname(base_title)}!", f"{quoted_new}!"),
    ]
    for old, new in replacements:
        text = text.replace(old, new)
    return text.encode("utf-8")

