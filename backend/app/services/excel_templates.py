from __future__ import annotations

import csv
import io
import re
import copy
from dataclasses import dataclass
from typing import Dict, List, Sequence
from xml.etree import ElementTree as ET
import zipfile
from copy import copy as clone_style

from openpyxl import load_workbook

_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XML_NS = "http://www.w3.org/XML/1998/namespace"
_XLSX_SHEET_PATH = "xl/worksheets/sheet1.xml"


@dataclass(frozen=True)
class ColumnSpec:
    key: str
    letter: str
    style: str


def _column_to_index(letter: str) -> int:
    result = 0
    for char in letter:
        if not char.isalpha():
            break
        result = result * 26 + (ord(char.upper()) - ord("A") + 1)
    return result


def _split_cell(reference: str) -> tuple[str, int]:
    match = re.match(r"([A-Z]+)(\d+)", reference)
    if not match:
        raise ValueError(f"셀 참조를 해석할 수 없습니다: {reference}")
    column, row = match.groups()
    return column, int(row)


def _parse_dimension(ref: str) -> tuple[str, int, str, int]:
    if ":" in ref:
        start_ref, end_ref = ref.split(":", 1)
    else:
        start_ref = end_ref = ref
    start_col, start_row = _split_cell(start_ref)
    end_col, end_row = _split_cell(end_ref)
    return start_col, start_row, end_col, end_row


class WorksheetPopulator:
    def __init__(
        self,
        sheet_bytes: bytes,
        *,
        start_row: int,
        columns: Sequence[ColumnSpec],
    ) -> None:
        self._ns = {"s": _SPREADSHEET_NS}
        self._root = ET.fromstring(sheet_bytes)
        self._sheet_data = self._root.find("s:sheetData", self._ns)
        if self._sheet_data is None:
            raise ValueError("워크시트 데이터 영역을 찾을 수 없습니다.")

        self._dimension = self._root.find("s:dimension", self._ns)
        if self._dimension is None:
            raise ValueError("워크시트 범위 정보를 찾을 수 없습니다.")
        ref = self._dimension.get("ref")
        if not ref:
            raise ValueError("워크시트 범위 정보를 확인할 수 없습니다.")
        (
            self._dimension_start_col,
            self._dimension_start_row,
            self._dimension_end_col,
            self._dimension_end_row,
        ) = _parse_dimension(ref)

        self._start_row = start_row
        self._column_specs = list(columns)
        if not self._column_specs:
            raise ValueError("채울 열 정보가 없습니다.")

        self._row_cache: Dict[int, ET.Element] = {}
        for row in self._sheet_data.findall("s:row", self._ns):
            r_attr = row.get("r")
            if not r_attr:
                continue
            try:
                index = int(r_attr)
            except ValueError:
                continue
            self._row_cache[index] = row

        template_row = self._row_cache.get(self._start_row)
        if template_row is None:
            raise ValueError("템플릿 행을 찾을 수 없습니다.")
        self._template_row = copy.deepcopy(template_row)

    def _tag(self, name: str) -> str:
        return f"{{{_SPREADSHEET_NS}}}{name}"

    @staticmethod
    def _cell_column(cell: ET.Element) -> str:
        ref = cell.get("r", "")
        return "".join(filter(str.isalpha, ref))

    def _ensure_row(self, index: int) -> ET.Element:
        if index in self._row_cache:
            return self._row_cache[index]

        row = copy.deepcopy(self._template_row)
        row.set("r", str(index))
        for cell in row.findall("s:c", self._ns):
            column = self._cell_column(cell)
            cell.set("r", f"{column}{index}")
            self._clear_cell(cell)
        self._sheet_data.append(row)
        self._row_cache[index] = row
        if index > self._dimension_end_row:
            self._dimension_end_row = index
        return row

    def _clear_cell(self, cell: ET.Element) -> None:
        if "t" in cell.attrib:
            del cell.attrib["t"]
        for child in list(cell):
            cell.remove(child)

    def _clear_row(self, row: ET.Element) -> None:
        for cell in row.findall("s:c", self._ns):
            self._clear_cell(cell)

    def _ensure_cell(self, row: ET.Element, spec: ColumnSpec) -> ET.Element:
        column = spec.letter
        target_index = _column_to_index(column)
        for cell in row.findall("s:c", self._ns):
            if self._cell_column(cell) == column:
                cell.set("r", f"{column}{row.get('r')}")
                cell.set("s", spec.style)
                return cell

        new_cell = ET.Element(self._tag("c"), {
            "r": f"{column}{row.get('r')}",
            "s": spec.style,
        })
        inserted = False
        for idx, existing in enumerate(list(row)):
            if existing.tag != self._tag("c"):
                continue
            existing_col = self._cell_column(existing)
            if _column_to_index(existing_col) > target_index:
                row.insert(idx, new_cell)
                inserted = True
                break
        if not inserted:
            row.append(new_cell)
        return new_cell

    def _set_cell_value(self, cell: ET.Element, value: str) -> None:
        self._clear_cell(cell)
        cleaned = value.strip()
        if not cleaned:
            return

        cell.set("t", "inlineStr")
        is_elem = ET.SubElement(cell, self._tag("is"))
        t_elem = ET.SubElement(is_elem, self._tag("t"))
        if cleaned != value or "\n" in value:
            t_elem.set(f"{{{_XML_NS}}}space", "preserve")
            t_elem.text = value
        else:
            t_elem.text = cleaned

    def populate(self, records: Sequence[Dict[str, str]]) -> None:
        # 우선 기존 데이터를 비웁니다.
        for index, row in self._row_cache.items():
            if index >= self._start_row:
                self._clear_row(row)

        for offset, record in enumerate(records):
            row_index = self._start_row + offset
            row = self._ensure_row(row_index)
            for spec in self._column_specs:
                value = record.get(spec.key, "")
                cell = self._ensure_cell(row, spec)
                self._set_cell_value(cell, value)

        limit = self._start_row + len(records)
        for index in sorted(self._row_cache):
            if index >= limit:
                row = self._row_cache[index]
                self._clear_row(row)

        if records:
            last_row = self._start_row + len(records) - 1
        else:
            last_row = self._start_row
        if last_row > self._dimension_end_row:
            self._dimension_end_row = last_row
        self._dimension.set(
            "ref",
            f"{self._dimension_start_col}{self._dimension_start_row}:{self._dimension_end_col}{self._dimension_end_row}",
        )

    def to_bytes(self) -> bytes:
        return ET.tostring(self._root, encoding="utf-8", xml_declaration=True)


def _replace_sheet_bytes(workbook_bytes: bytes, new_sheet_bytes: bytes) -> bytes:
    source_buffer = io.BytesIO(workbook_bytes)
    output_buffer = io.BytesIO()
    with zipfile.ZipFile(source_buffer, "r") as source_zip:
        with zipfile.ZipFile(output_buffer, "w") as target_zip:
            for info in source_zip.infolist():
                data = source_zip.read(info.filename)
                if info.filename == _XLSX_SHEET_PATH:
                    data = new_sheet_bytes
                target_zip.writestr(info, data)
    return output_buffer.getvalue()


def _parse_csv_records(csv_text: str, expected_columns: Sequence[str]) -> List[Dict[str, str]]:
    stripped = csv_text.strip()
    if not stripped:
        return []

    reader = csv.reader(io.StringIO(stripped))
    rows = [row for row in reader]
    if not rows:
        return []

    header = [cell.strip() for cell in rows[0]]
    if header:
        header[0] = header[0].lstrip("\ufeff")
    column_index: Dict[str, int] = {}
    for idx, name in enumerate(header):
        if name:
            column_index[name] = idx

    missing = [column for column in expected_columns if column not in column_index]
    if missing:
        raise ValueError(f"CSV에 필요한 열이 없습니다: {', '.join(missing)}")

    records: List[Dict[str, str]] = []
    for raw in rows[1:]:
        entry: Dict[str, str] = {}
        is_empty = True
        for column in expected_columns:
            idx = column_index[column]
            value = raw[idx].strip() if idx < len(raw) else ""
            if value:
                is_empty = False
            entry[column] = value
        if not is_empty:
            records.append(entry)
    return records


FEATURE_LIST_COLUMNS: Sequence[ColumnSpec] = (
    ColumnSpec(key="대분류", letter="A", style="12"),
    ColumnSpec(key="중분류", letter="B", style="8"),
    ColumnSpec(key="소분류", letter="C", style="15"),
)

FEATURE_LIST_EXPECTED_HEADERS: Sequence[str] = ["대분류", "중분류", "소분류"]


def populate_feature_list(workbook_bytes: bytes, csv_text: str) -> bytes:
    records = _parse_csv_records(csv_text, FEATURE_LIST_EXPECTED_HEADERS)
    with zipfile.ZipFile(io.BytesIO(workbook_bytes), "r") as source:
        sheet_bytes = source.read(_XLSX_SHEET_PATH)
    populator = WorksheetPopulator(sheet_bytes, start_row=8, columns=FEATURE_LIST_COLUMNS)
    populator.populate(records)
    return _replace_sheet_bytes(workbook_bytes, populator.to_bytes())


TESTCASE_COLUMNS: Sequence[ColumnSpec] = (
    ColumnSpec(key="대분류", letter="A", style="31"),
    ColumnSpec(key="중분류", letter="B", style="31"),
    ColumnSpec(key="소분류", letter="C", style="18"),
    ColumnSpec(key="테스트 케이스 ID", letter="D", style="24"),
    ColumnSpec(key="테스트 시나리오", letter="E", style="18"),
    ColumnSpec(key="입력(사전조건 포함)", letter="F", style="18"),
    ColumnSpec(key="기대 출력(사후조건 포함)", letter="G", style="18"),
    ColumnSpec(key="테스트 결과", letter="H", style="19"),
    ColumnSpec(key="상세 테스트 결과", letter="I", style="7"),
    ColumnSpec(key="비고", letter="J", style="6"),
)

TESTCASE_EXPECTED_HEADERS: Sequence[str] = [
    "대분류",
    "중분류",
    "소분류",
    "테스트 케이스 ID",
    "테스트 시나리오",
    "입력(사전조건 포함)",
    "기대 출력(사후조건 포함)",
    "테스트 결과",
    "상세 테스트 결과",
    "비고",
]


def populate_testcase_list(workbook_bytes: bytes, csv_text: str) -> bytes:
    records = _parse_csv_records(csv_text, TESTCASE_EXPECTED_HEADERS)
    with zipfile.ZipFile(io.BytesIO(workbook_bytes), "r") as source:
        sheet_bytes = source.read(_XLSX_SHEET_PATH)
    populator = WorksheetPopulator(sheet_bytes, start_row=6, columns=TESTCASE_COLUMNS)
    populator.populate(records)
    return _replace_sheet_bytes(workbook_bytes, populator.to_bytes())


SECURITY_REPORT_COLUMNS: Sequence[ColumnSpec] = (
    ColumnSpec(key="순번", letter="A", style="24"),
    ColumnSpec(key="시험환경 OS", letter="B", style="25"),
    ColumnSpec(key="결함 요약", letter="C", style="10"),
    ColumnSpec(key="결함 정도", letter="D", style="26"),
    ColumnSpec(key="발생 빈도", letter="E", style="26"),
    ColumnSpec(key="품질 특성", letter="F", style="25"),
    ColumnSpec(key="결함 설명", letter="G", style="23"),
    ColumnSpec(key="업체 응답", letter="H", style="10"),
    ColumnSpec(key="수정여부", letter="I", style="10"),
    ColumnSpec(key="비고", letter="J", style="11"),
)

SECURITY_REPORT_EXPECTED_HEADERS: Sequence[str] = [
    "순번",
    "시험환경 OS",
    "결함 요약",
    "결함 정도",
    "발생 빈도",
    "품질 특성",
    "결함 설명",
    "업체 응답",
    "수정여부",
    "비고",
]


def populate_security_report(workbook_bytes: bytes, csv_text: str) -> bytes:
    records = _parse_csv_records(csv_text, SECURITY_REPORT_EXPECTED_HEADERS)
    if not records:
        return workbook_bytes

    workbook = load_workbook(io.BytesIO(workbook_bytes))
    worksheet = workbook.active

    start_row = 6

    # Determine existing populated rows by inspecting 결함 요약 열 (C)
    current_row = start_row
    while True:
        cell_value = worksheet.cell(row=current_row, column=_column_to_index("C")).value
        if cell_value is None or str(cell_value).strip() == "":
            break
        current_row += 1

    existing_count = current_row - start_row

    template_cells = {
        spec.letter: worksheet[f"{spec.letter}{start_row}"] for spec in SECURITY_REPORT_COLUMNS
    }

    for index, record in enumerate(records, start=1):
        row_index = current_row + index - 1
        record_index = existing_count + index
        normalized_record = dict(record)
        normalized_record["순번"] = str(record_index)

        for spec in SECURITY_REPORT_COLUMNS:
            cell = worksheet[f"{spec.letter}{row_index}"]
            template_cell = template_cells.get(spec.letter)
            if template_cell is not None and template_cell.has_style:
                cell.font = clone_style(template_cell.font)
                cell.fill = clone_style(template_cell.fill)
                cell.border = clone_style(template_cell.border)
                cell.alignment = clone_style(template_cell.alignment)
                cell.number_format = template_cell.number_format
                cell.protection = clone_style(template_cell.protection)

            cell.value = normalized_record.get(spec.key, "")

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
