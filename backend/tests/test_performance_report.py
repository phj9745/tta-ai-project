from __future__ import annotations

import io
import zipfile
from pathlib import Path
from typing import Any, Dict, Optional, Sequence
from xml.etree import ElementTree as ET

import pytest
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from app.services.performance_report import (
    TEMPLATE_PATH,
    PerformanceReportBuilder,
    PerformanceReportService,
    detect_os,
    parse_raw_dataset,
)
from app.services.google_drive import XLSX_MIME_TYPE

SAMPLE_ROOT = Path("backend/samplefiles")
WINDOWS_RAW = SAMPLE_ROOT / "rawdata" / "DataCollector01.csv"
LINUX_RAW = SAMPLE_ROOT / "rawdata" / "linux.txt"


def _load_bytes(path: Path) -> bytes:
    return path.read_bytes()


def _namespace() -> Dict[str, str]:
    return {"c": "http://schemas.openxmlformats.org/drawingml/2006/chart"}


def test_detect_os_from_content() -> None:
    windows_text = r"\\Processor(_Total)\% Processor Time"
    linux_text = "procs -----------memory---------- ---swap--"

    assert detect_os("perf.csv", windows_text).value == "Windows"
    assert detect_os("vmstat.txt", linux_text).value == "Linux"


def test_builder_generates_workbook_with_updated_charts() -> None:
    template_bytes = _load_bytes(TEMPLATE_PATH)
    windows_dataset = parse_raw_dataset("DataCollector01.csv", _load_bytes(WINDOWS_RAW))
    linux_dataset = parse_raw_dataset("linux.txt", _load_bytes(LINUX_RAW))

    builder = PerformanceReportBuilder(template_bytes)
    output_bytes = builder.build([windows_dataset, linux_dataset])

    # Verify sheet names and inserted values.
    workbook = load_workbook(io.BytesIO(output_bytes), data_only=False)
    assert workbook.sheetnames == ["시나리오", "Windows #1", "Linux #1", "응답시간"]

    filter_db = next(
        (dn for dn in workbook.defined_names.definedName if dn.name == "_xlnm._FilterDatabase"),
        None,
    )
    assert filter_db is not None
    assert filter_db.attr_text.startswith("'Windows #1'!")

    windows_sheet = workbook["Windows #1"]
    first_win_record = windows_dataset.records[0]
    assert windows_sheet["D4"].value == first_win_record["timestamp"]
    assert windows_sheet["E4"].value == first_win_record["cpu_percent"]
    assert windows_sheet["F4"].value == first_win_record["private_bytes"]

    linux_sheet = workbook["Linux #1"]
    first_linux_record = linux_dataset.records[0]
    assert linux_sheet["D4"].value == first_linux_record["us"]
    assert linux_sheet["E4"].value == first_linux_record["sy"]
    assert linux_sheet["F4"].value == first_linux_record["free"]
    assert linux_sheet["G4"].value == first_linux_record["buff"]
    assert linux_sheet["H4"].value == first_linux_record["cache"]

    win_last_row = PerformanceReportBuilder.START_ROW + len(windows_dataset.records) - 1
    linux_last_row = PerformanceReportBuilder.START_ROW + len(linux_dataset.records) - 1

    with zipfile.ZipFile(io.BytesIO(output_bytes)) as archive:
        chart1 = ET.fromstring(archive.read("xl/charts/chart1.xml"))
        chart7 = ET.fromstring(archive.read("xl/charts/chart7.xml"))
        assert not any(name.startswith("xl/externalLinks/") for name in archive.namelist())
        chart1_xml = ET.tostring(chart1, encoding="utf-8").decode("utf-8")
        chart7_xml = ET.tostring(chart7, encoding="utf-8").decode("utf-8")
        assert "Windows!" not in chart1_xml
        assert "Linux!" not in chart7_xml
    ns = _namespace()

    cat_range = chart1.find(".//c:cat/c:numRef/c:f", ns)
    val_range = chart1.find(".//c:val/c:numRef/c:f", ns)
    assert cat_range is not None
    assert val_range is not None
    assert cat_range.text == f"'Windows #1'!$A$4:$A${win_last_row}"
    assert val_range.text == f"'Windows #1'!$H$4:$H${win_last_row}"

    linux_cat = chart7.find(".//c:cat/c:numRef/c:f", ns)
    linux_val = chart7.find(".//c:val/c:numRef/c:f", ns)
    assert linux_cat is not None
    assert linux_val is not None
    assert linux_cat.text == f"'Linux #1'!$A$4:$A${linux_last_row}"
    assert linux_val.text == f"'Linux #1'!$L$4:$L${linux_last_row}"


class _DummyDriveService:
    def __init__(self, template_bytes: bytes) -> None:
        self._template = template_bytes
        self.upload_payload: Optional[Dict[str, Any]] = None
        self.created_paths: Optional[Sequence[str]] = None

    async def get_project_exam_number(self, *, project_id: str, google_id: Optional[str]) -> str:
        return "GS-X-99-9999"

    async def upload_file_to_folder(
        self,
        *,
        parent_id: str,
        file_name: str,
        content: bytes,
        google_id: Optional[str],
        content_type: str = XLSX_MIME_TYPE,
        ) -> Dict[str, Any]:
        self.upload_payload = {
            "parent_id": parent_id,
            "file_name": file_name,
            "content": content,
            "content_type": content_type,
            "google_id": google_id,
        }
        return {"id": "uploaded-file"}

    async def ensure_project_subfolder(
        self,
        *,
        project_id: str,
        path: Sequence[str],
        google_id: Optional[str],
    ) -> str:
        self.created_paths = tuple(path)
        return f"{project_id}-{'/'.join(path)}"


@pytest.mark.asyncio
async def test_performance_report_service_flow() -> None:
    template_bytes = _load_bytes(TEMPLATE_PATH)
    drive_service = _DummyDriveService(template_bytes)
    service = PerformanceReportService(drive_service)  # type: ignore[arg-type]

    upload = UploadFile(filename="DataCollector01.csv", file=io.BytesIO(_load_bytes(WINDOWS_RAW)))
    workbook = await service.generate_report(
        project_id="dummy-project",
        uploads=[upload],
        google_id="user-123",
    )

    assert workbook.filename == "GS-X-99-9999 성능시험.xlsx"
    assert isinstance(workbook.content, bytes)
    assert drive_service.upload_payload is not None
    assert drive_service.upload_payload["file_name"] == workbook.filename
    assert drive_service.upload_payload["parent_id"] == "dummy-project-다.수행/성능시험"
    assert drive_service.upload_payload["content_type"] == XLSX_MIME_TYPE
    assert drive_service.created_paths == ("다.수행", "성능시험")


@pytest.mark.asyncio
async def test_generate_report_requires_os_selection() -> None:
    template_bytes = _load_bytes(TEMPLATE_PATH)
    drive_service = _DummyDriveService(template_bytes)
    service = PerformanceReportService(drive_service)  # type: ignore[arg-type]

    ambiguous_csv = (
        "\"Timestamp\",\"Private Bytes\",\"% Processor Time\"\\n"
        "\"2025-01-01 00:00:00\",\"1024\",\"12.5\"\\n"
    ).encode("utf-8")

    upload = UploadFile(filename="manual.csv", file=io.BytesIO(ambiguous_csv))

    with pytest.raises(HTTPException) as thrown:
        await service.generate_report(
            project_id="dummy-project",
            uploads=[upload],
            google_id=None,
        )

    detail = thrown.value.detail
    assert isinstance(detail, dict)
    assert detail.get("code") == "os_selection_required"
    assert detail.get("files") == [{"filename": "manual.csv", "index": 0}]

    second_upload = UploadFile(filename="manual.csv", file=io.BytesIO(ambiguous_csv))
    workbook = await service.generate_report(
        project_id="dummy-project",
        uploads=[second_upload],
        google_id=None,
        os_hints=["Windows"],
    )

    assert workbook.filename == "GS-X-99-9999 성능시험.xlsx"
    assert drive_service.upload_payload is not None
