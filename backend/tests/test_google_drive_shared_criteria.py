from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")
from openpyxl import load_workbook

# Ensure backend/app is importable when executing tests from repo root.
BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.services.google_drive.templates import (  # noqa: E402
    build_default_shared_criteria_workbook,
    is_shared_criteria_candidate,
    replace_placeholders,
)


def test_build_default_shared_criteria_workbook_has_expected_headers() -> None:
    """Fallback workbook must include the required shared criteria headers."""
    workbook_bytes = build_default_shared_criteria_workbook()
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    workbook.close()
    assert headers == [
        "Invicti 결과",
        "결함 요약",
        "결함정도",
        "발생빈도",
        "품질특성",
        "결함 설명",
        "결함 제외 여부",
    ]


def test_is_shared_criteria_candidate_handles_whitespace() -> None:
    assert is_shared_criteria_candidate("  결함판단 기준표 v1.0 .xlsx  ")
    assert not is_shared_criteria_candidate("other.xlsx")


def test_replace_placeholders_replaces_all_known_tokens() -> None:
    text = "GS-B-XX-XXXX vs GS-B-2X-XXXX"
    assert replace_placeholders(text, "GS-X-99-9999") == "GS-X-99-9999 vs GS-X-99-9999"
